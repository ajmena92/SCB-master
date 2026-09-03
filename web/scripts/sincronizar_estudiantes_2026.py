#!/usr/bin/env python3
"""Sincroniza de forma cerrada el padrón REGULAR y transporte 2026 desde Excel.

La simulación no requiere base. Con DATABASE_URL también incluye el plan contra
PostgreSQL; sólo ``--aplicar`` escribe y todas las escrituras van en una transacción.
La columna ``matricula.becado`` nunca se inserta ni se actualiza en conflictos.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import secrets
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from argon2 import PasswordHasher
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

sys.path.insert(0, str(Path(__file__).parent))
from importar_sqlserver_postgresql import _codigo, _texto

HOJA_PADRON = "query"
HOJA_RUTAS = "rptListadoEstudiantes"
ENCABEZADOS_PADRON = {"Cédula", "Title", "Segundo Apellido", "Nombre", "Sección", "Estado"}
ENCABEZADOS_RUTAS = {"Ruta", "Identificación", "Nombre", "Estado"}
def normalizar_cedula(valor: Any) -> str | None:
    """Devuelve una identificación no vacía, sin suponer un tipo documental.

    Las identificaciones son texto: se conserva cada letra y dígito, incluidos
    ceros iniciales y documentos alfanuméricos o extranjeros. Sólo se eliminan
    espacios y guiones de presentación para conciliar formatos equivalentes; no
    se rellenan, truncan ni corrigen dígitos.
    """
    original = unicodedata.normalize("NFKC", _texto(valor)).strip()
    if not original:
        return None
    return re.sub(r"[\s-]+", "", original).upper()


def _fila_error(tipo: str, fila: dict[str, Any], **extra: Any) -> dict[str, Any]:
    return {"tipo": tipo, "fila": fila["fila"], **extra}


def _leer_hoja(archivo: Path, hoja_esperada: str, requeridas: set[str]) -> list[dict[str, Any]]:
    libro = load_workbook(archivo, read_only=True, data_only=True)
    if libro.sheetnames != [hoja_esperada]:
        raise ValueError(f"{archivo.name}: se esperaba únicamente la hoja {hoja_esperada!r}")
    hoja = libro[hoja_esperada]
    filas = hoja.iter_rows(values_only=True)
    try:
        encabezados = [_texto(valor) for valor in next(filas)]
    except StopIteration as error:
        raise ValueError(f"{archivo.name}: hoja vacía") from error
    faltantes = requeridas - set(encabezados)
    if faltantes:
        raise ValueError(f"{archivo.name}: faltan columnas requeridas: {', '.join(sorted(faltantes))}")
    resultado = []
    for numero, valores in enumerate(filas, start=2):
        if any(valor is not None and _texto(valor) for valor in valores):
            resultado.append(dict(zip(encabezados, valores, strict=True)) | {"fila": numero})
    return resultado


def leer_padron(archivo: Path) -> list[dict[str, Any]]:
    return [{"fila": f["fila"], "cedula": f["Cédula"],
             "nombre": _texto(" ".join(filter(None, [_texto(f["Nombre"]), _texto(f["Title"]), _texto(f["Segundo Apellido"])]))),
             "seccion": _texto(f["Sección"]), "estado": _texto(f["Estado"])}
            for f in _leer_hoja(archivo, HOJA_PADRON, ENCABEZADOS_PADRON)]


def leer_rutas(archivo: Path) -> list[dict[str, Any]]:
    return [{"fila": f["fila"], "cedula": f["Identificación"], "ruta": _texto(f["Ruta"]),
             "nombre": _texto(f["Nombre"]), "estado": _texto(f["Estado"])}
            for f in _leer_hoja(archivo, HOJA_RUTAS, ENCABEZADOS_RUTAS)]


def validar_padron(filas: list[dict[str, Any]]) -> tuple[dict[str, dict[str, str]], list[dict[str, Any]]]:
    regulares, errores, vistos = {}, [], Counter()
    for fila in filas:
        if fila["estado"].upper() == "REGULAR":
            cedula = normalizar_cedula(fila["cedula"])
            if cedula: vistos[cedula] += 1
    for fila in filas:
        if fila["estado"].upper() != "REGULAR":
            continue
        cedula = normalizar_cedula(fila["cedula"])
        if not cedula: errores.append(_fila_error("cedula_padron_invalida", fila)); continue
        if vistos[cedula] > 1: errores.append(_fila_error("cedula_padron_duplicada", fila, cedula=cedula)); continue
        if not fila["nombre"]: errores.append(_fila_error("nombre_padron_ausente", fila)); continue
        if not fila["seccion"]: errores.append(_fila_error("seccion_padron_ausente", fila)); continue
        regulares[cedula] = {"cedula": cedula, "nombres": fila["nombre"], "seccion": fila["seccion"]}
    return regulares, errores


def validar_rutas(
    filas: list[dict[str, Any]], cedulas_regulares: set[str]
) -> tuple[dict[str, str], list[dict[str, Any]], list[dict[str, Any]]]:
    """Valida rutas del universo REGULAR y deja trazas de las filas omitidas.

    Las rutas de personas fuera del padrón REGULAR no son errores de la fuente
    primaria: se excluyen, aun si sus estados son contradictorios, y se reportan
    como advertencias para conciliación posterior.
    """
    errores, advertencias, por_cedula = [], [], defaultdict(list)
    for fila in filas:
        cedula = normalizar_cedula(fila["cedula"])
        if not cedula: errores.append(_fila_error("cedula_ruta_invalida", fila)); continue
        estado = fila["estado"].casefold()
        por_cedula[cedula].append((estado, fila))
    activas: dict[str, str] = {}
    for cedula, entradas in por_cedula.items():
        if cedula not in cedulas_regulares:
            advertencias.extend(_fila_error("ruta_para_estudiante_no_regular", fila, cedula=cedula) for _, fila in entradas)
            estados = {estado for estado, _ in entradas}
            if {"activo", "disminuido"}.issubset(estados):
                advertencias.extend(_fila_error("estado_ruta_contradictorio", fila, cedula=cedula) for _, fila in entradas)
            continue
        entradas = [(estado, fila) for estado, fila in entradas if estado in {"activo", "disminuido"}]
        if not entradas:
            continue
        for _, fila in entradas:
            if not fila["ruta"]:
                errores.append(_fila_error("ruta_ausente", fila, cedula=cedula))
        estados = {estado for estado, _ in entradas}
        if estados == {"activo", "disminuido"}:
            errores.extend(_fila_error("estado_ruta_contradictorio", fila, cedula=cedula) for _, fila in entradas); continue
        activas_filas = [fila for estado, fila in entradas if estado == "activo"]
        if len({fila["ruta"] for fila in activas_filas}) > 1 or len(activas_filas) > 1:
            errores.extend(_fila_error("ruta_activa_duplicada", fila, cedula=cedula) for fila in activas_filas); continue
        if activas_filas and not any(not fila["ruta"] for fila in activas_filas):
            fila = activas_filas[0]
            activas[cedula] = fila["ruta"]
    return activas, errores, advertencias


def matriculas_sin_ruta_activa(matriculas: dict[str, int], rutas_activas: dict[str, str]) -> list[int]:
    """Matrículas REGULAR cuya asignación vigente debe cerrarse por no estar Activo."""
    return [matricula_id for cedula, matricula_id in matriculas.items() if cedula not in rutas_activas]


def fecha_cierre_ruta(fecha_efectiva: date) -> date:
    """El día de término es inclusivo; evita dos rutas vigentes el mismo día."""
    return fecha_efectiva - timedelta(days=1)


def debe_desactivar_persona(cedula: str, cedulas_regulares: set[str]) -> bool:
    """Desactiva sólo identificaciones no vacías ausentes del padrón REGULAR."""
    normalizada = normalizar_cedula(cedula)
    return normalizada is not None and normalizada not in cedulas_regulares


def plan_rutas(actuales: dict[str, dict[str, Any]], rutas_fuente: dict[str, str]) -> dict[str, int]:
    """Resume altas, cambios y cierres sin exponer cédulas fuera del detalle controlado."""
    crear = cambiar = cerrar = 0
    for cedula, ruta in rutas_fuente.items():
        actual = actuales.get(cedula, {}).get("ruta_codigo")
        if actual is None: crear += 1
        elif actual not in {ruta, f"RUTA-{ruta}"}: cambiar += 1
    for cedula, actual in actuales.items():
        if actual.get("ruta_codigo") is not None and cedula not in rutas_fuente: cerrar += 1
    return {"rutas_a_crear": crear, "rutas_a_cambiar": cambiar, "rutas_a_cerrar": cerrar}


def _fuente(archivo: Path) -> dict[str, str]:
    return {"archivo": str(archivo), "sha256": hashlib.sha256(archivo.read_bytes()).hexdigest()}


def escribir_credenciales(archivo: Path, credenciales: list[tuple[str, str, str]]) -> None:
    """Escribe los PIN nuevos tras confirmar la transacción, con permisos 0600.

    El archivo es deliberadamente independiente del reporte JSON: los PIN son
    secretos de entrega y nunca deben llegar a la consola ni al registro de la
    importación. ``os.open`` fija los permisos al crear el archivo, sin depender
    de la umask del proceso.
    """
    archivo.parent.mkdir(parents=True, exist_ok=True)
    banderas = os.O_WRONLY | os.O_CREAT | os.O_TRUNC
    if hasattr(os, "O_NOFOLLOW"):
        banderas |= os.O_NOFOLLOW
    descriptor = os.open(archivo, banderas, 0o600)
    os.fchmod(descriptor, 0o600)
    with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as salida:
        escritor = csv.writer(salida)
        escritor.writerow(["codigo", "cedula", "pin_temporal"])
        escritor.writerows(credenciales)


def plan_personas_fuente(
    padron: dict[str, dict[str, str]], personas_actuales: list[dict[str, Any]]
) -> tuple[dict[str, dict[str, Any]], list[str], list[str], list[str], list[dict[str, Any]]]:
    """Concilia el padrón contra todas las personas, no sólo un año lectivo."""
    por_cedula: dict[str, dict[str, Any]] = {}
    errores: list[dict[str, Any]] = []
    for persona in personas_actuales:
        cedula = normalizar_cedula(persona["cedula"])
        if not cedula:
            errores.append({"tipo": "cedula_base_invalida", "persona_id": persona["persona_id"]})
            continue
        if cedula in por_cedula:
            errores.append({"tipo": "cedula_base_duplicada", "cedula": cedula})
            continue
        por_cedula[cedula] = persona
    crear = sorted(set(padron) - set(por_cedula))
    actualizar = sorted(set(padron) & set(por_cedula))
    desactivar = sorted(
        cedula for cedula, persona in por_cedula.items()
        if persona["activo"] and cedula not in padron
    )
    return por_cedula, crear, actualizar, desactivar, errores


def planificar(url: str, anio: int, padron: dict[str, dict[str, str]], rutas: dict[str, str]) -> dict[str, Any]:
    motor = create_engine(url, pool_pre_ping=True)
    with motor.connect() as conexion:
        actuales = conexion.execute(text("""SELECT p.id persona_id,p.cedula,p.nombres,p.activo,m.id matricula_id,
          m.seccion,m.turno,m.estado,m.becado,a.ruta_id,r.codigo ruta_codigo
          FROM persona p LEFT JOIN matricula m ON m.persona_id=p.id
          LEFT JOIN anio_lectivo al ON al.id=m.anio_lectivo_id
          LEFT JOIN asignacion_ruta a ON a.matricula_id=m.id AND a.fecha_fin IS NULL
          LEFT JOIN ruta r ON r.id=a.ruta_id WHERE p.tipo='estudiante' AND al.anio=:anio"""), {"anio": anio}).mappings().all()
        todas_personas = conexion.execute(text("""SELECT id persona_id,cedula,activo
          FROM persona WHERE tipo='estudiante'""")).mappings().all()
    por_cedula, errores = {}, []
    for actual in actuales:
        cedula = normalizar_cedula(actual["cedula"])
        if not cedula: errores.append({"tipo": "cedula_base_invalida", "persona_id": actual["persona_id"]}); continue
        if cedula in por_cedula: errores.append({"tipo": "cedula_base_duplicada", "cedula": cedula}); continue
        por_cedula[cedula] = actual
    _, crear, actualizar, desactivar_personas, errores_personas = plan_personas_fuente(padron, todas_personas)
    errores.extend(errores_personas)
    desactivar_matriculas = sorted(set(por_cedula) - set(padron))
    # aplicar() desactiva personas de cualquier año; el preview debe presentar
    # exactamente ese alcance y no sólo las matrículas del año seleccionado.
    rutas_plan = plan_rutas(por_cedula, rutas)
    return {"personas_crear": len(crear), "personas_actualizar": len(actualizar), "personas_desactivar": len(desactivar_personas), "matriculas_desactivar": len(desactivar_matriculas),
            "rutas_activas_fuente": len(rutas), "becas_comedor_preservadas": True, "errores": errores,
            "detalle": {"cedulas_crear": crear, "cedulas_desactivar": desactivar_personas,
                        "cedulas_matriculas_desactivar": desactivar_matriculas}, **rutas_plan}


def aplicar(
    url: str,
    anio: int,
    padron: dict[str, dict[str, str]],
    rutas: dict[str, str],
    semilla: str,
    credenciales_salida: Path,
) -> dict[str, int]:
    if len(semilla) < 32: raise ValueError("CODIGO_MIGRACION_SEMILLA debe tener al menos 32 caracteres")
    motor, conteos = create_engine(url, pool_pre_ping=True), Counter()
    credenciales_nuevas: list[tuple[str, str, str]] = []
    fecha_efectiva = date.today()
    fecha_cierre = fecha_cierre_ruta(fecha_efectiva)
    with motor.begin() as c:
        c.execute(text("SELECT pg_advisory_xact_lock(:bloqueo)"), {"bloqueo": anio})
        anio_id = c.execute(text("INSERT INTO anio_lectivo(anio,vigente) VALUES (:anio,false) ON CONFLICT(anio) DO UPDATE SET anio=excluded.anio RETURNING id"), {"anio": anio}).scalar_one()
        existentes = {normalizar_cedula(x["cedula"]): x for x in c.execute(text("SELECT p.id,p.cedula,m.id matricula_id FROM persona p LEFT JOIN matricula m ON m.persona_id=p.id AND m.anio_lectivo_id=:anio WHERE p.tipo='estudiante'"), {"anio": anio_id}).mappings() if normalizar_cedula(x["cedula"])}
        todas_personas = c.execute(text("SELECT id,cedula FROM persona WHERE tipo='estudiante'")) .mappings().all()
        hasher = PasswordHasher()
        matriculas = {}
        for indice, (cedula, fila) in enumerate(sorted(padron.items()), 1):
            actual = existentes.get(cedula)
            if actual:
                persona_id = actual["id"]
                c.execute(text("UPDATE persona SET nombres=:nombres,activo=true WHERE id=:id"), fila | {"id": persona_id})
            else:
                codigo = _codigo("estudiante", anio * 100000 + indice, semilla + cedula)
                persona_id = c.execute(text("INSERT INTO persona(codigo,cedula,nombres,tipo,activo) VALUES (:codigo,:cedula,:nombres,'estudiante',true) RETURNING id"), fila | {"codigo": codigo}).scalar_one()
                pin_temporal = f"{secrets.randbelow(1_000_000):06d}"
                c.execute(text("INSERT INTO credencial_portal(persona_id,pin_hash,cambio_obligatorio) VALUES (:id,:hash,true)"), {"id": persona_id, "hash": hasher.hash(pin_temporal)})
                c.execute(text("INSERT INTO cuenta_tiquete(persona_id,saldo,reservados) VALUES (:id,0,0)"), {"id": persona_id})
                credenciales_nuevas.append((codigo, cedula, pin_temporal))
                conteos["personas_creadas"] += 1
            matriculas[cedula] = c.execute(text("""INSERT INTO matricula(persona_id,anio_lectivo_id,seccion,turno,becado,estado)
              VALUES (:persona,:anio,:seccion,'diurno',false,'activo') ON CONFLICT(persona_id,anio_lectivo_id)
              DO UPDATE SET seccion=excluded.seccion,turno='diurno',estado='activo' RETURNING id"""), fila | {"persona": persona_id, "anio": anio_id}).scalar_one()
        for cedula, actual in existentes.items():
            if cedula not in padron and actual["matricula_id"]:
                cambio = c.execute(text("UPDATE matricula SET estado='trasladado' WHERE id=:id AND estado='activo'"), {"id": actual["matricula_id"]}); conteos["matriculas_trasladadas"] += cambio.rowcount
        for persona in todas_personas:
            if debe_desactivar_persona(persona["cedula"], set(padron)):
                cambio = c.execute(text("UPDATE persona SET activo=false WHERE id=:id AND activo=true"), {"id": persona["id"]})
                if cambio.rowcount:
                    conteos["personas_desactivadas"] += 1
                    conteos["sesiones_revocadas"] += c.execute(text("DELETE FROM sesion_acceso WHERE persona_id=:id"), {"id": persona["id"]}).rowcount
        for cedula, codigo_origen in rutas.items():
            codigo = f"RUTA-{codigo_origen}"
            ruta = c.execute(text("SELECT id,activo FROM ruta WHERE codigo IN (:codigo,:origen)"), {"codigo": codigo, "origen": codigo_origen}).mappings().first()
            ruta_id = ruta["id"] if ruta else None
            if ruta_id is None:
                ruta_id = c.execute(text("INSERT INTO ruta(nombre,codigo,descripcion,color_hex,activo) VALUES (:nombre,:codigo,:descripcion,'#CBD5E1',true) RETURNING id"), {"nombre": f"Ruta {codigo_origen}", "codigo": codigo, "descripcion": "Importada desde rutas_estudiants.xlsx"}).scalar_one(); conteos["rutas_creadas"] += 1
            elif not ruta["activo"]:
                c.execute(text("UPDATE ruta SET activo=true WHERE id=:id"), {"id": ruta_id}); conteos["rutas_reactivadas"] += 1
            c.execute(text("UPDATE asignacion_ruta SET fecha_fin=:fin WHERE matricula_id=:m AND fecha_fin IS NULL AND ruta_id<>:r"), {"fin": fecha_cierre, "m": matriculas[cedula], "r": ruta_id})
            existe = c.execute(text("SELECT 1 FROM asignacion_ruta WHERE matricula_id=:m AND fecha_fin IS NULL"), {"m": matriculas[cedula]}).scalar()
            if not existe: c.execute(text("INSERT INTO asignacion_ruta(matricula_id,ruta_id,fecha_inicio,fecha_fin) VALUES (:m,:r,:inicio,NULL)"), {"m": matriculas[cedula], "r": ruta_id, "inicio": fecha_efectiva}); conteos["rutas_asignadas"] += 1
        for matricula_id in matriculas_sin_ruta_activa(matriculas, rutas):
            cierre = c.execute(text("UPDATE asignacion_ruta SET fecha_fin=:fin WHERE matricula_id=:m AND fecha_fin IS NULL"), {"fin": fecha_cierre, "m": matricula_id})
            conteos["rutas_cerradas"] += cierre.rowcount
        c.execute(text("UPDATE asignacion_ruta SET fecha_fin=:fin WHERE fecha_fin IS NULL AND matricula_id IN (SELECT m.id FROM matricula m JOIN anio_lectivo al ON al.id=m.anio_lectivo_id JOIN persona p ON p.id=m.persona_id WHERE al.anio=:anio AND p.tipo='estudiante' AND m.estado<>'activo')"), {"fin": fecha_cierre, "anio": anio})
    # El contexto anterior ya confirmó PostgreSQL. Si la transacción falla, esta
    # línea no se alcanza y no se deja ningún PIN temporal en disco.
    escribir_credenciales(credenciales_salida, credenciales_nuevas)
    return dict(conteos)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("padron", type=Path); p.add_argument("rutas", type=Path); p.add_argument("--anio", type=int, default=2026)
    directorio_privado = Path(os.getenv("REPORTE_PRIVADO_DIR", "/tmp"))
    p.add_argument("--aplicar", action="store_true"); p.add_argument("--reporte", type=Path, default=directorio_privado / "reporte-sincronizacion-2026.json")
    p.add_argument("--credenciales", type=Path, default=directorio_privado / "credenciales-sincronizacion-2026.csv",
                   help="CSV confidencial para PIN nuevos; se crea sólo tras confirmar PostgreSQL")
    args = p.parse_args()
    try:
        padron, e_padron = validar_padron(leer_padron(args.padron)); rutas, e_rutas, advertencias = validar_rutas(leer_rutas(args.rutas), set(padron))
        reporte: dict[str, Any] = {"modo": "aplicar" if args.aplicar else "simulacion", "anio": args.anio, "fuentes": [_fuente(args.padron), _fuente(args.rutas)], "estudiantes_regulares": len(padron), "rutas_activas": len(rutas), "errores": e_padron + e_rutas, "advertencias": advertencias, "becas_comedor": "preservadas sin cambios"}
        url = os.getenv("DATABASE_URL")
        if not reporte["errores"] and url: reporte["plan_base"] = planificar(url, args.anio, padron, rutas); reporte["errores"].extend(reporte["plan_base"]["errores"])
        if args.aplicar:
            if reporte["errores"]:
                reporte["aplicado"] = False
                reporte["mensaje"] = "Hay errores bloqueantes; no se aplicó ninguna modificación. Corrija la fuente usando el detalle del reporte."
            elif not url:
                raise ValueError("DATABASE_URL es requerida con --aplicar")
            else:
                reporte["aplicado"] = aplicar(url, args.anio, padron, rutas, os.getenv("CODIGO_MIGRACION_SEMILLA", ""), args.credenciales)
        codigo = 2 if reporte["errores"] else 0
    except (ValueError, OSError) as error:
        reporte, codigo = {"modo": "aplicar" if args.aplicar else "simulacion", "error": str(error)}, 2
    args.reporte.parent.mkdir(parents=True, exist_ok=True)
    args.reporte.write_text(json.dumps(reporte, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    try: args.reporte.chmod(0o600)
    except OSError: pass
    print(json.dumps(reporte, ensure_ascii=False, indent=2)); return codigo


if __name__ == "__main__": raise SystemExit(main())
