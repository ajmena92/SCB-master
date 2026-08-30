#!/usr/bin/env python3
"""Vista previa y aplicación idempotente del padrón anual XLSX."""

from __future__ import annotations

import argparse
import csv
import json
import os
import secrets
import sys
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

from argon2 import PasswordHasher
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from importar_sqlserver_postgresql import _codigo, _texto


REQUERIDAS = {"cedula", "nombres", "tipo"}


def _clave(valor: Any) -> str:
    normal = unicodedata.normalize("NFKD", _texto(valor)).encode("ascii", "ignore").decode()
    return normal.lower().replace(" ", "_")


def leer(archivo: Path) -> list[dict[str, Any]]:
    libro = load_workbook(archivo, read_only=True, data_only=True)
    hoja = libro.active
    filas = hoja.iter_rows(values_only=True)
    encabezados = [_clave(valor) for valor in next(filas)]
    faltantes = REQUERIDAS - set(encabezados)
    if faltantes:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(sorted(faltantes))}")
    return [dict(zip(encabezados, valores, strict=True)) | {"fila": indice}
            for indice, valores in enumerate(filas, start=2)
            if any(valor is not None and _texto(valor) for valor in valores)]


def normalizar(filas: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    cedulas = Counter(_texto(fila.get("cedula")) for fila in filas if _texto(fila.get("cedula")))
    validas, errores = [], []
    for fila in filas:
        cedula = _texto(fila.get("cedula"))
        tipo = _clave(fila.get("tipo"))
        nombres = _texto(fila.get("nombres"))
        problemas = []
        if not cedula: problemas.append("cedula_ausente")
        if cedula and cedulas[cedula] > 1: problemas.append("cedula_duplicada")
        if not nombres: problemas.append("nombre_ausente")
        if tipo not in {"estudiante", "profesor"}: problemas.append("tipo_invalido")
        if tipo == "estudiante" and not _texto(fila.get("seccion")): problemas.append("seccion_ausente")
        if tipo == "estudiante" and not _texto(fila.get("turno")): problemas.append("turno_ausente")
        if problemas:
            errores.append({"fila": fila["fila"], "errores": problemas})
            continue
        validas.append({
            "fila": fila["fila"], "cedula": cedula, "nombres": nombres, "tipo": tipo,
            "seccion": _texto(fila.get("seccion")) or None,
            "turno": _texto(fila.get("turno")) or None,
            "becado": _clave(fila.get("becado")) in {"si", "true", "1", "becado"},
            "ruta": _texto(fila.get("ruta")) or None,
            "estado": _clave(fila.get("estado")) or "activo",
        })
    return validas, errores


def aplicar(url: str, anio: int, filas: list[dict[str, Any]], semilla: str, salida: Path) -> dict[str, int]:
    motor, hasher, conteos, credenciales = create_engine(url, pool_pre_ping=True), PasswordHasher(), Counter(), []
    with motor.begin() as conexion:
        anio_id = conexion.execute(text(
            """INSERT INTO anio_lectivo(anio,vigente) VALUES (:anio,false)
            ON CONFLICT(anio) DO UPDATE SET anio=excluded.anio RETURNING id"""
        ), {"anio": anio}).scalar_one()
        for indice, fila in enumerate(filas, start=1):
            existente = conexion.execute(text("SELECT id,codigo,tipo FROM persona WHERE cedula=:cedula"), fila).mappings().first()
            if existente:
                if existente["tipo"] != fila["tipo"]:
                    raise ValueError(f"La fila {fila['fila']} cambia el tipo de una persona existente")
                persona_id, codigo = existente["id"], existente["codigo"]
                conexion.execute(text(
                    "UPDATE persona SET nombres=:nombres,tipo=:tipo,activo=(:estado='activo') WHERE id=:id"
                ), fila | {"id": persona_id})
                conteos["personas_actualizadas"] += 1
            else:
                codigo = _codigo(fila["tipo"], anio * 100_000 + indice, semilla + fila["cedula"])
                persona_id = conexion.execute(text(
                    """INSERT INTO persona(codigo,cedula,nombres,tipo,activo)
                    VALUES (:codigo,:cedula,:nombres,:tipo,(:estado='activo')) RETURNING id"""
                ), fila | {"codigo": codigo}).scalar_one()
                pin = f"{secrets.randbelow(1_000_000):06d}"
                conexion.execute(text(
                    "INSERT INTO credencial_portal(persona_id,pin_hash,cambio_obligatorio) VALUES (:id,:hash,true)"
                ), {"id": persona_id, "hash": hasher.hash(pin)})
                credenciales.append((codigo, fila["cedula"], pin))
                conteos["personas_creadas"] += 1
            if fila["tipo"] != "estudiante":
                continue
            matricula_id = conexion.execute(text(
                """INSERT INTO matricula(persona_id,anio_lectivo_id,seccion,turno,becado,estado)
                VALUES (:persona,:anio,:seccion,:turno,:becado,:estado)
                ON CONFLICT(persona_id,anio_lectivo_id) DO UPDATE SET seccion=excluded.seccion,
                turno=excluded.turno,becado=excluded.becado,estado=excluded.estado RETURNING id"""
            ), fila | {"persona": persona_id, "anio": anio_id}).scalar_one()
            if fila["ruta"]:
                ruta_id = conexion.execute(text(
                    """INSERT INTO ruta(nombre,activo) VALUES (:ruta,true)
                    ON CONFLICT(nombre) DO UPDATE SET activo=true RETURNING id"""
                ), fila).scalar_one()
                parametros_ruta = {"matricula": matricula_id, "ruta": ruta_id, "anio": anio}
                actualizado = conexion.execute(text(
                    "UPDATE asignacion_ruta SET ruta_id=:ruta WHERE matricula_id=:matricula AND fecha_fin IS NULL"
                ), parametros_ruta)
                if actualizado.rowcount == 0:
                    conexion.execute(text(
                        """INSERT INTO asignacion_ruta(matricula_id,ruta_id,fecha_inicio,fecha_fin)
                        VALUES (:matricula,:ruta,make_date(:anio,1,1),NULL)"""
                    ), parametros_ruta)
            conteos["matriculas"] += 1
    salida.parent.mkdir(parents=True, exist_ok=True)
    with salida.open("w", encoding="utf-8", newline="") as archivo:
        escritor = csv.writer(archivo); escritor.writerow(["codigo", "cedula", "pin_temporal"]); escritor.writerows(credenciales)
    try: salida.chmod(0o600)
    except OSError: pass
    return dict(conteos)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("archivo", type=Path)
    parser.add_argument("--anio", type=int, required=True)
    parser.add_argument("--aplicar", action="store_true")
    parser.add_argument("--reporte", type=Path, default=Path("reporte-padron.json"))
    parser.add_argument("--credenciales", type=Path, default=Path("credenciales-padron.csv"))
    args = parser.parse_args()
    filas, errores = normalizar(leer(args.archivo))
    reporte: dict[str, Any] = {"modo": "aplicar" if args.aplicar else "simulacion", "anio": args.anio,
                               "filas_validas": len(filas), "errores": errores}
    resultado = 0 if not errores else 2
    if args.aplicar:
        if errores:
            print("Hay errores bloqueantes; no se aplicó el padrón.", file=sys.stderr)
        else:
            url, semilla = os.getenv("DATABASE_URL", ""), os.getenv("CODIGO_MIGRACION_SEMILLA", "")
            if not url or len(semilla) < 32: parser.error("DATABASE_URL y CODIGO_MIGRACION_SEMILLA son requeridas")
            reporte["aplicados"] = aplicar(url, args.anio, filas, semilla, args.credenciales)
    args.reporte.write_text(json.dumps(reporte, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(reporte, ensure_ascii=False, indent=2))
    return resultado


if __name__ == "__main__": raise SystemExit(main())
