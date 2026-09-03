"""Previsualizacion, lectura XLSX y confirmacion idempotente."""

import hashlib
import io
import json
import secrets
from typing import Literal, cast

from fastapi import HTTPException

from aplicacion.esquemas import FilaImportacion, ImportacionEntrada
from aplicacion.modelos.maestros import AnioLectivo, Matricula, Persona
from aplicacion.modelos.operacion import LoteImportacion
from aplicacion.seguridad import hash_secreto


class ServicioImportacion:
    def __init__(self, repo):
        self.repo = repo

    def desde_excel(self, contenido: bytes, anio: int) -> ImportacionEntrada:
        try:
            from openpyxl import load_workbook

            hoja = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True).active
            filas = list(hoja.iter_rows(values_only=True))
            encabezados = [str(v).strip().lower() if v else "" for v in filas[0]]
            requeridas = {"cedula", "nombres", "tipo"}
            if not requeridas <= set(encabezados):
                raise ValueError("faltan columnas cedula, nombres o tipo")
            datos = []
            for valores in filas[1:]:
                fila = dict(zip(encabezados, valores))
                if not any(v is not None for v in valores):
                    continue
                tipo = str(fila.get("tipo") or "").lower()
                datos.append(
                    FilaImportacion(
                        cedula=str(fila.get("cedula") or "") or None,
                        nombres=str(fila.get("nombres") or ""),
                        tipo=cast(Literal["estudiante", "profesor"], tipo),
                        seccion=str(fila.get("seccion") or "") or None,
                    )
                )
            return ImportacionEntrada(anio=anio, filas=datos)
        except Exception as exc:
            raise HTTPException(422, f"Excel invalido: {exc}") from exc

    def _huella(self, datos):
        return hashlib.sha256(
            json.dumps(datos.model_dump(mode="json"), sort_keys=True).encode()
        ).hexdigest()

    def previsualizar(self, datos):
        errores = []
        cedulas = set()
        altas = cambios = 0
        tipos = set()
        for indice, fila in enumerate(datos.filas, 1):
            if not fila.cedula:
                errores.append({"fila": indice, "error": "cedula requerida"})
                continue
            if fila.cedula in cedulas:
                errores.append({"fila": indice, "error": "cedula duplicada"})
                continue
            cedulas.add(fila.cedula)
            tipos.add(fila.tipo)
            existente = self.repo.persona_cedula(fila.cedula)
            altas += existente is None
            cambios += existente is not None
            if existente and existente.tipo != fila.tipo:
                errores.append({"fila": indice, "error": "el tipo no coincide con la persona existente"})
            if fila.tipo == "estudiante" and not fila.seccion:
                errores.append({"fila": indice, "error": "seccion requerida"})
        desactivaciones = len(self.repo.activas_ausentes_del_padron(tipos, cedulas))
        return {
            "huella": self._huella(datos),
            "total": len(datos.filas),
            "altas": altas,
            "cambios": cambios,
            "desactivaciones": desactivaciones,
            "errores": errores,
            "aplicable": not errores,
        }

    def confirmar(self, datos, huella):
        resumen = self.previsualizar(datos)
        if resumen["huella"] != huella:
            raise HTTPException(409, "El contenido cambio")
        if not resumen["aplicable"]:
            raise HTTPException(422, detail=resumen["errores"])
        lote = self.repo.lote(huella)
        if lote:
            return {"loteId": lote.id, "repetida": True, "credenciales": [], **resumen}
        anio = self.repo.anio(datos.anio) or self.repo.guardar(
            AnioLectivo(anio=datos.anio, vigente=False)
        )
        credenciales = []
        tipos = {fila.tipo for fila in datos.filas}
        cedulas = {fila.cedula for fila in datos.filas if fila.cedula}
        ausentes = self.repo.activas_ausentes_del_padron(tipos, cedulas)
        for fila in datos.filas:
            persona = self.repo.persona_cedula(fila.cedula)
            if not persona:
                persona = Persona(
                    cedula=fila.cedula.strip(),
                    nombres=fila.nombres,
                    tipo=fila.tipo,
                    activo=True,
                )
                pin_temporal = f"{secrets.randbelow(1_000_000):06d}"
                self.repo.guardar_persona_nueva(persona, hash_secreto(pin_temporal))
                credenciales.append(
                    {
                        "cedula": persona.cedula,
                        "nombre": persona.nombres,
                        "pinTemporal": pin_temporal,
                    }
                )
            else:
                persona.nombres, persona.activo = fila.nombres, True
            if fila.tipo != "estudiante":
                continue
            matricula = self.repo.matricula(persona.id, anio.id) or Matricula(
                persona_id=persona.id, anio_lectivo_id=anio.id
            )
            matricula.seccion, matricula.turno, matricula.estado = (
                fila.seccion or "",
                "diurno",
                "activo",
            )
            self.repo.guardar(matricula)
        self.repo.desactivar_personas(ausentes)
        lote = self.repo.guardar(
            LoteImportacion(huella=huella, estado="confirmado", resumen=json.dumps(resumen))
        )
        return {
            "loteId": lote.id,
            "repetida": False,
            "credenciales": credenciales,
            **resumen,
        }
